"""Excel processing module for reading XLSX files and QuickBooks integration.

This module provides functions to read Excel files, specifically payment terms,
and integrate with QuickBooks Desktop via COM API.
"""

import xml.etree.ElementTree as ET
from dataclasses import dataclass
from typing import Any
from xml.sax.saxutils import escape

import win32com.client
from openpyxl import load_workbook


@dataclass
class PaymentTerm:
    """Represents a payment term with name and discount days."""

    name: str
    discount_days: int


def read_payment_terms(file_path: str) -> list[PaymentTerm]:
    """Read payment terms from the specified Excel file.

    Expected Excel format:
    - Must contain a sheet named 'payment_terms'
    - Column A: Payment term names (strings)
    - Column B: Discount days (integers)
    - Row 1 should contain headers (will be skipped)
    - Data starts from row 2

    Args:
        file_path (str): Path to the Excel file containing payment terms (.xlsx format)

    Returns:
        list[PaymentTerm]: List of payment terms with name and discount days.
                          Empty list if no valid payment terms found.

    Raises:
        No exceptions need to be manually raised - let openpyxl handle file/sheet errors

    Implementation Notes:
        - Use openpyxl.load_workbook() with read_only=True for better performance
        - Access the 'payment_terms' worksheet by name
        - Use worksheet.iter_rows(min_row=2, values_only=True) to skip headers
        - Validate that both name (column A) and discount_days (column B) are not None
        - Convert name to string and strip whitespace
        - Convert discount_days to integer, skip rows with invalid data
        - Handle ValueError/TypeError when converting discount_days to int
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb["payment_terms"]
    except KeyError:
        # Let openpyxl/caller handle missing-sheet semantics; re-raise for clarity
        wb.close()
        raise

    payment_terms: list[PaymentTerm] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # row may be a tuple like (name, discount_days, ...)
        if not row:
            continue

        name_cell = row[0] if len(row) >= 1 else None
        days_cell = row[1] if len(row) >= 2 else None

        # Skip rows where name is missing or blank
        if name_cell is None:
            continue
        name = str(name_cell).strip()
        if not name:
            continue

        # Convert discount days to int, skip if invalid
        if days_cell is None:
            continue
        try:
            discount_days = int(days_cell)
        except (TypeError, ValueError):
            try:
                discount_days = int(float(days_cell))
            except Exception as e:
                raise TypeError(f"Invalid discount_days for term '{name}': {days_cell!r}") from e

        payment_terms.append(PaymentTerm(name=name, discount_days=discount_days))

    wb.close()
    return payment_terms


def connect_to_quickbooks() -> Any:
    """Connect to QuickBooks Desktop via COM API.

    This function establishes a connection to QuickBooks Desktop using the
    QBXML Request Processor COM interface. QuickBooks Desktop must be running
    with a company file open.

    Returns:
        tuple[Any, Any]: A tuple containing (qb_app, session)
            - qb_app: COM object for QuickBooks application interface
            - session: Session ticket for the current QB connection

    Raises:
        No exceptions need to be manually raised - let win32com handle COM errors

    Security Notes:
        - User may need to grant permission in QuickBooks for first-time access
        - QuickBooks may prompt user to allow external application access
    """
    try:
        qb_app = win32com.client.Dispatch("QBXMLRP2.RequestProcessor")
        qb_app.OpenConnection("", "Payment Terms Import")
        session = qb_app.BeginSession("", 2)  # 2 = qbFileOpenDoNotCare
        return qb_app, session
    except Exception as e:
        print(f"QuickBooks connection error: {str(e)}")
        raise


def create_payment_terms_batch_qbxml(payment_terms: list[PaymentTerm]) -> str:
    """Create QBXML for adding multiple payment terms in a batch.

    This function generates a well-formed QBXML document containing multiple
    StandardTermsAddRq requests that can be sent to QuickBooks Desktop in a
    single batch operation.

    Args:
        payment_terms (list[PaymentTerm]): List of payment terms to create.
                                         Each PaymentTerm must have name and discount_days.

    Returns:
        str: Complete QBXML string ready to send to QuickBooks Desktop.
             Contains XML declaration, QBXML root, and multiple StandardTermsAddRq elements.

    Raises:
        AttributeError: If PaymentTerm objects are missing required attributes
        TypeError: If payment_terms is not a list or contains invalid objects

    QBXML Structure:
        <?xml version="1.0" encoding="utf-8"?>
        <?qbxml version="13.0"?>
        <QBXML>
            <QBXMLMsgsRq onError="continueOnError">
                <StandardTermsAddRq>
                    <StandardTermsAdd>
                        <Name>Term Name</Name>
                        <StdDueDays>Number of Days</StdDueDays>
                    </StandardTermsAdd>
                </StandardTermsAddRq>
                ... (repeated for each payment term)
            </QBXMLMsgsRq>
        </QBXML>

    Implementation Notes:
        - Create a list to store individual StandardTermsAddRq XML strings
        - Loop through payment_terms and create XML for each term
        - Use f-strings to format XML with term.name and term.discount_days
        - XML escape special characters in term names if necessary
        - Join all requests with newlines using chr(10).join()
        - Wrap in proper QBXML envelope with version="13.0"
        - Use onError="continueOnError" to process all terms even if some fail
        - Note: <StdDueDays > has trailing space - this is required by QB format
    """
    if not isinstance(payment_terms, list):
        raise TypeError("payment_terms must be a list of PaymentTerm objects")

    requests: list[str] = []

    for term in payment_terms:
        if not hasattr(term, "name") or not hasattr(term, "discount_days"):
            raise AttributeError(
                "Each item in payment_terms must have 'name' and 'discount_days' attributes"
            )

        name = str(term.name).strip()
        if not name:
            # skip empty names
            continue

        # ensure discount_days is an integer
        try:
            discount_days = int(term.discount_days)
        except (TypeError, ValueError):
            try:
                discount_days = int(float(term.discount_days))
            except Exception as err:
                raise TypeError(
                    f"Invalid discount_days for term '{name}': {term.discount_days!r}"
                ) from err

        # Escape XML special characters in the term name
        safe_name = escape(name)

        # Note: <StdDueDays > has a trailing space by QB format expectation
        rq = (
            "<StandardTermsAddRq>"
            "<StandardTermsAdd>"
            f"<Name>{safe_name}</Name>"
            f"<StdDueDays >{discount_days}</StdDueDays >"
            "</StandardTermsAdd>"
            "</StandardTermsAddRq>"
        )
        requests.append(rq)

    inner = chr(10).join(requests)

    qbxml = (
        '<?xml version="1.0" encoding="utf-8"?>'
        + chr(10)
        + '<?qbxml version="13.0"?>'
        + chr(10)
        + "<QBXML>"
        + chr(10)
        + '<QBXMLMsgsRq onError="continueOnError">'
        + chr(10)
        + inner
        + (chr(10) if inner else "")
        + "</QBXMLMsgsRq>"
        + chr(10)
        + "</QBXML>"
    )

    return qbxml


def save_payment_terms_to_quickbooks(payment_terms: list[PaymentTerm]) -> list[str]:
    """Save payment terms to QuickBooks Desktop.

    This function connects to QuickBooks, sends a batch QBXML request to create
    multiple payment terms, parses the response, and returns the list of
    successfully created terms.

    Args:
        payment_terms (list[PaymentTerm]): List of payment terms to save to QuickBooks.
                                         Each term must have valid name and discount_days.

    Returns:
        list[str]: List of payment term names that were successfully created.
                  May be shorter than input list if some terms failed or already exist.

    Raises:
        RuntimeError: If connection to QuickBooks fails (manually wrap exceptions for clarity)

    Response Parsing Logic:
        - Success: statusCode="0" indicates successful creation
        - Already exists: Error code "3100" means term already exists (skip silently)
        - Other errors: Print warning and exclude from results

    Implementation Notes:
        - Call connect_to_quickbooks() to establish QB connection
        - Use create_payment_terms_batch_qbxml() to generate QBXML request
        - Send QBXML using qb_app.ProcessRequest(session, qbxml)
        - Parse XML response to identify successful operations:
            * Look for statusCode="0" AND term name in response
            * Handle "3100" error code (term already exists) gracefully
            * Print warnings for other failures
        - Always cleanup: call qb_app.EndSession(session) and qb_app.CloseConnection()
        - Use try/except to catch and re-raise exceptions with meaningful messages
        - Consider that QB response may contain multiple StandardTermsAddRs elements

    QuickBooks Error Codes:
        - 0: Success
        - 3100: Object already exists
        - Other codes indicate various QB-specific errors
    """
    if not isinstance(payment_terms, list):
        raise TypeError("payment_terms must be a list of PaymentTerm objects")

    if not payment_terms:
        return []

    qbxml = create_payment_terms_batch_qbxml(payment_terms)

    qb_app = None
    session = None
    created: list[str] = []

    try:
        qb_app, session = connect_to_quickbooks()
        # send request
        response = qb_app.ProcessRequest(session, qbxml)

        # response should be an XML string; try to parse
        try:
            root = ET.fromstring(response)
        except Exception:
            # fall back to parsing substrings if XML parsing fails
            # attempt simple extraction of <Name>...</Name> entries for statusCode="0"
            created = []
            if 'statusCode="0"' in response:
                # naive: collect Names in order
                for part in response.split("<Name>")[1:]:
                    name = part.split("</Name>", 1)[0]
                    created.append(name)
            return created

        # find all StandardTermsAddRs elements and inspect statusCode
        for rs in root.findall(".//StandardTermsAddRs"):
            status = rs.attrib.get("statusCode")
            # If success, try to locate the Name in the returned StandardTermsAddRet
            if status == "0":
                name_el = rs.find(".//Name")
                if name_el is not None and name_el.text:
                    created.append(name_el.text)
                else:
                    # If Name not present where expected, try to infer from request order:
                    # skip — this is rare; continue
                    continue
            elif status == "3100":
                # Already exists — skip silently
                continue
            else:
                # Other errors: print a warning with available info
                msg_el = rs.find(".//StatusMessage")
                msg = msg_el.text if (msg_el is not None and msg_el.text) else None
                print(f"QuickBooks returned error for a term (status {status}): {msg}")

        return created
    except Exception as exc:
        # wrap common COM errors for caller clarity
        raise RuntimeError(f"Failed to connect to QuickBooks: {exc}") from exc
    finally:
        try:
            if qb_app is not None and session is not None:
                try:
                    qb_app.EndSession(session)
                except Exception:
                    pass
                try:
                    qb_app.CloseConnection()
                except Exception:
                    pass
        except Exception:
            pass


def process_payment_terms(file_path: str) -> list[str]:
    """Read payment terms from Excel and save to QuickBooks.

    This is the main orchestration function that handles the complete workflow:
    reading payment terms from an Excel file and saving them to QuickBooks Desktop.

    Args:
        file_path (str): Path to the Excel file containing payment terms (.xlsx format).
                        File must contain a 'payment_terms' sheet with Name and ID columns.

    Returns:
        list[str]: List of payment term names that were successfully created in QuickBooks.
                  May be empty if no terms were processed or all failed.

    Raises:
        ValueError: If no payment terms found in the Excel file (manually check and raise)

    Workflow:
        1. Read payment terms from Excel file using read_payment_terms()
        2. Validate that at least one payment term was found
        3. Print summary of terms to be imported (for user feedback)
        4. Save all terms to QuickBooks using save_payment_terms_to_quickbooks()

    Implementation Notes:
        - Call read_payment_terms(file_path) to extract payment terms from Excel
        - Check if payment_terms list is empty and raise ValueError with helpful message
        - Print found terms for user visibility before QB operation
        - Use f"Found {len(payment_terms)} payment terms to import:" for count
        - Print each term as f"  - {term.name} ({term.discount_days} days)"
        - Let QuickBooks connection errors bubble up naturally from save function
        - No need for separate QB connection test - let save_payment_terms handle it

    Error Handling Strategy:
        - Validate Excel data before attempting QuickBooks operations
        - Provide clear error messages for common issues
        - Let underlying functions handle their specific error cases
        - Don't catch and re-wrap exceptions unless adding meaningful context
    """
    payment_terms = read_payment_terms(file_path)

    if not payment_terms:
        raise ValueError("No payment terms found in the 'payment_terms' sheet.")

    print(f"Found {len(payment_terms)} payment terms to import:")
    for term in payment_terms:
        print(f"  - {term.name} ({term.discount_days} days)")

    # attempt to save to QuickBooks and return created names
    return save_payment_terms_to_quickbooks(payment_terms)
